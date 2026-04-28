import sys
import os
os.environ['PYTHONUNBUFFERED'] = '1'

from flask import Flask, render_template, request, send_file
import fitz
import re
import json
import base64
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import datetime

app = Flask(__name__)
UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

MISTRAL_API_KEY = os.environ.get("MISTRAL_API_KEY")

def pdf_to_base64_images(filepath):
    doc = fitz.open(filepath)
    images = []
    mat = fitz.Matrix(2, 2)
    for page in doc:
        pix = page.get_pixmap(matrix=mat, alpha=False)
        img_bytes = pix.tobytes("png")
        b64 = base64.b64encode(img_bytes).decode("utf-8")
        images.append(b64)
    return images

def image_to_base64(filepath):
    with open(filepath, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")

def ocr_with_mistral(filepath):
    prompt = """Analizá este comprobante bancario o documento financiero y extraé los datos en formato JSON.
Si no encontrás algún campo dejalo como string vacío.
Devolvé SOLO el JSON sin texto adicional ni markdown.

{
  "tipo_documento": "",
  "banco": "",
  "fecha": "",
  "hora": "",
  "importe": "",
  "moneda": "",
  "cuenta_origen": "",
  "titular_origen": "",
  "cuit_origen": "",
  "cuenta_destino": "",
  "titular_destino": "",
  "cuit_destino": "",
  "nro_referencia": "",
  "motivo": "",
  "concepto": "",
  "estado": ""
}"""

    if filepath.lower().endswith(".pdf"):
        images_b64 = pdf_to_base64_images(filepath)
        if not images_b64:
            return {}
        b64 = images_b64[0]
        media_type = "image/png"
    else:
        b64 = image_to_base64(filepath)
        ext = filepath.lower().split(".")[-1]
        media_type = "image/jpeg" if ext in ["jpg", "jpeg"] else "image/png"

    response = requests.post(
        "https://api.mistral.ai/v1/chat/completions",
        headers={
            "Authorization": f"Bearer {MISTRAL_API_KEY}",
            "Content-Type": "application/json"
        },
        json={
            "model": "pixtral-12b-latest",
            "messages": [
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image_url",
                            "image_url": f"data:{media_type};base64,{b64}"
                        },
                        {
                            "type": "text",
                            "text": prompt
                        }
                    ]
                }
            ]
        }
    )

    try:
        text = response.json()["choices"][0]["message"]["content"].strip()
        text = re.sub(r"```json|```", "", text).strip()
        return json.loads(text)
    except Exception as e:
        sys.stdout.write(f"ERROR PARSEANDO: {e}\n")
        sys.stdout.flush()
        return {}

def generate_excel(results):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comprobantes"

    headers = ["Archivo", "Tipo", "Banco", "Fecha", "Hora", "Importe", "Moneda",
               "Cuenta Origen", "Titular Origen", "CUIT Origen",
               "Cuenta Destino", "Titular Destino", "CUIT Destino",
               "N° Referencia", "Motivo", "Concepto", "Estado"]

    header_fill = PatternFill(start_color="1a3d1a", end_color="1a3d1a", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 20

    for result in results:
        data = result["data"]
        row = [
            result["filename"],
            data.get("tipo_documento", ""),
            data.get("banco", ""),
            data.get("fecha", ""),
            data.get("hora", ""),
            data.get("importe", ""),
            data.get("moneda", ""),
            data.get("cuenta_origen", ""),
            data.get("titular_origen", ""),
            data.get("cuit_origen", ""),
            data.get("cuenta_destino", ""),
            data.get("titular_destino", ""),
            data.get("cuit_destino", ""),
            data.get("nro_referencia", ""),
            data.get("motivo", ""),
            data.get("concepto", ""),
            data.get("estado", "")
        ]
        ws.append(row)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

@app.route("/", methods=["GET", "POST"])
def index():
    results = []
    if request.method == "POST":
        files = request.files.getlist("file")
        for file in files:
            if file and file.filename:
                filename = file.filename
                filepath = os.path.join(UPLOAD_FOLDER, filename)
                file.save(filepath)
                data = ocr_with_mistral(filepath)
                parsed = [{"Campo": k.replace("_", " ").title(), "Valor": v} for k, v in data.items()]
                results.append({"filename": filename, "parsed": parsed, "data": data})
    return render_template("index.html", results=results)

@app.route("/descargar", methods=["POST"])
def descargar():
    data_json = request.form.get("data")
    results = json.loads(data_json)
    buffer = generate_excel(results)
    filename = f"comprobantes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(buffer, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    app.run(host="0.0.0.0", port=port, debug=False)